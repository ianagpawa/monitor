import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
import datetime

# Find start and end rows of invoice
def find_start(document_name):
    wb = openpyxl.load_workbook(document_name)
    sheet = wb.get_sheet_by_name('Sheet1')
    print 'Finding start row of invoices'
    for i in range(1, 125):
        location = 'A%s' % i
        cell = sheet[location]
        value = cell.internal_value
        if value == 1:
            start = i
        if ("Number" in str(value)) and i > 1:
            last = i - 1
            print "Finished getting start and last rows"
            return start, last

def get_invoice_number(document_name):
    words = document_name.split(" ")
    return words[1]


def check_name(document_name, start):
    start -= 1
    wb = openpyxl.load_workbook(document_name)
    sheet = wb.get_sheet_by_name('Sheet1')
    for i in range(1,15):
        cell = sheet.cell(row=start, column=i)
        if 'Therapist' in cell.value:
            return get_column_letter(i)


def get_values(document_name):
    data = {}

    wb = openpyxl.load_workbook(document_name)
    sheet = wb.get_sheet_by_name('Sheet1')
    start, last = find_start(document_name)

    therapist_c = check_name(document_name, start)

    for i in range(start, last):
        date = sheet["B" + str(i)].value
        site = sheet["C" + str(i)].value
        therapist = sheet[therapist_c + str(i)].value
        if site in data:
            arr = data[site]["date"]
            if date not in arr:
                arr.append(date)
        else:
            data[site] = {"therapist": therapist, "date": [date]}
    return data
    # results = open('test.py', 'w')
    # results.write('allData = ' + pprint.pformat(data))
    # results.close()
    # print ("Done.")


# Create new calander sheet
def create_calendar(month):
    wb = openpyxl.load_workbook("calendar.xlsx")
    print 'Creating new sheet'
    new_month_sheet = wb.create_sheet(title=month)
    new_month_sheet.page_setup.orientation = new_month_sheet.ORIENTATION_LANDSCAPE


    first_cell = new_month_sheet.cell(row=1, column=1)
    first_cell.value = month

    print "Creating days"
    for i in range(2, 33):
        cell = new_month_sheet.cell(row=2, column=i)
        column = cell.column

        new_month_sheet.column_dimensions[column].width = "15"
        cell.value = i - 1
    wb.save('calendar.xlsx')
    print "Done.  Created new sheet for " + month
    with open("log.txt", "a") as text_file:
        text_file.write("Done.  Created new sheet for %s \n" % month)

# Add info from parsed invoice, single data obj, to calendar
def add_site(wb, sheet, site):
    for i in range(3, 50):
        cell = sheet.cell(row=i, column=1)
        if cell.value:
            if cell.value == site:
                return i
        else:
            row = sheet.row_dimensions[i]
            row.height = "60"

            cell.value = site
            print 'Adding %s to %s' % (site, sheet.title)
            with open("log.txt", "a") as text_file:
                text_file.write('Adding %s to %s \n' % (site, sheet.title))
            wb.save('calendar.xlsx')
            return i

def get_initials(name):
    names = name.split(" ")
    finished = ''
    for name in names:
        finished += name[0]
    return finished


colors = {
    "OT": "E7C80E",
    "SP": "004C00",
    "Psy": "660000"
}


def add_to_calendar(site, obj, invoice_num, color):
    dates = obj['date']
    therapist = obj['therapist']
    initials = get_initials(therapist)

    color_fill = PatternFill(start_color= color, end_color= color, fill_type="solid")


    wb = openpyxl.load_workbook("calendar.xlsx")

    error = []
    for da in dates:
        sheet_names = wb.sheetnames
        month = da.strftime("%B")
        day = int(da.strftime("%d"))

        if month not in sheet_names:
            create_calendar(month)
            wb = openpyxl.load_workbook("calendar.xlsx")

        current_sheet = wb[month]
        current_row = add_site(wb, current_sheet, site)
        cell = current_sheet.cell(row=current_row, column=day+1)
        if cell.value:
            if initials in cell.value:
                print 'ERROR: %s already submmited for %s %s at %s!' % (therapist, month, day, site)
                with open("log.txt", "a") as text_file:
                    text_file.write('ERROR: %s already submmited for %s %s at %s! \n' % (therapist, month, day, site))
                with open("doubles.txt", "a") as doubles_text:
                    doubles_text.write('ERROR: %s already submmited for %s %s at %s! \n' % (therapist, month, day, site))
                error.append((therapist, month, day, site))
                return

            else:

                cell.value += " and " + ("%s-%s" % (invoice_num, initials))
                cell.font = Font(color=color)
                with open("same_day.txt", "a") as same_day:
                    same_day.write('%s and another therapist submitted for the same day, %s %s at %s \n' % (therapist, month, day, site))

        else:
            cell.font = Font(color='FFFFFF')
            cell.value = "%s-%s" % (invoice_num, initials)
            cell.fill = color_fill



        # if error:
        #     return
        wb.save('calendar.xlsx')
        print "Added %s %s" % (da.strftime("%B %d"), initials)
        with open("log.txt", "a") as text_file:
            text_file.write("Added %s %s \n" % (da.strftime("%B %d"), initials))



def add_invoice(document_name):
    invoice_num = get_invoice_number(document_name)

    print "Starting %s" % invoice_num
    with open("log.txt", "a") as text_file:
        text_file.write("\n%s \nStarting %s \n" % (datetime.datetime.now(), invoice_num))

    modality = document_name[-8:-5].strip()
    color = colors[modality]

    invoice_data = get_values(document_name)
    sites = invoice_data.keys()
    for site in sites:
        site_obj = invoice_data[site]
        add_to_calendar(site, site_obj, invoice_num, color)
    print "Finished with %s" % invoice_num
    with open("log.txt", "a") as text_file:
        text_file.write("Finished with %s \n%s \n \n" % (invoice_num, datetime.datetime.now()))
