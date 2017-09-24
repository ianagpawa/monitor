import openpyxl
import os

colors = {
    "E7C80E": "OT",
    "004C00": "SP",
    "660000": "Psy"
}


def check_calendar():
    wb = openpyxl.load_workbook('calendar.xlsx')
    print 'Opening calendar'
    for sheet in wb.worksheets:
        pass


def check_month(worksheet):
    i = 3
    cell = worksheet.cell(row=i, column=1)
    for row in worksheet.iter_rows(min_row=2, max_col=32, max_row=40):
        if row[0].value:
            check_site(row)



def check_site(row):
    site = row[0]
    for cell in row[2:]:
        if cell.value:
            # print cell.coordinate, cell.value
            initials, modality = get_values(cell)
            print initials, modality


def get_values(cell):
    background_color = cell.fill.fgColor.index[2:]
    modality = colors[background_color]

    initials = cell.value.split("-")[1]

    return initials, modality

wb = openpyxl.load_workbook('calendar.xlsx')
sheet = wb['June']
check_month(sheet)
